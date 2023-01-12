import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.core.utils import ChromeType
# from webdriver_manager.microsoft import EdgeChromiumDriverManager
# from webdriver_manager.opera import OperaDriverManager

import os
import smtplib
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def read_excel():
    reader = pd.read_excel('test_case/merolagani_testcase.xlsx')
    for row, column in reader.iterrows():
        sn = column["SN"]
        test_summary = column["Test_Summary"]
        xpath = column["Xpath"]
        action = column["Action"]
        value = column["Value"]
        action_defination(sn, test_summary, xpath, action, value)


def action_defination(sn, test_summary, xpath, action, value):
    if action == 'open_browser':
        result, remarks = open_browser(value)
    elif action == 'open_url':
        result, remarks = open_url(value)
    elif action == 'click':
        result, remarks = click(xpath)
    elif action == 'hover':
        result, remarks = hover(xpath, value)
    elif action == 'date_picker':
        result, remarks = date_picker(xpath, value)
    # elif action == 'alert':
    #     result, remarks = alert(value)
    elif action == 'compare_text':
        result, remarks = compare_text(xpath, value)
    elif action == 'input_text':
        result, remarks = input_text(xpath, value)
    elif action == 'new_tab':
        result, remarks = new_tab(xpath)
    elif action == 'select_dropdown':
        result, remarks = select_dropdown(xpath, value)
    elif action == 'close_browser':
        result, remarks = close_browser()
    elif action == 'wait':
        result, remarks = wait(value)
    else:
        result = "FAIL"
        remarks = (action, "Not Supported")
    print(sn, test_summary, result, remarks)
    action_defination(sn, test_summary, xpath, action, value)

    excel_operation.write_result(sn, test_summary, result, remarks)


def write_result(sn, test_summary, result, remarks):
    workbook = openpyxl.load_workbook('Test_Result/Automation_Test_Result.xlsx')
    worksheet2 = workbook.get_sheet_by_name('Details')
    row = int(sn) + 1
    worksheet2.cell(row, 1, sn)
    worksheet2.cell(row, 2, test_summary)
    worksheet2.cell(row, 3, result)
    worksheet2.cell(row, 4, str(remarks))
    workbook.save('Test_Result/Automation_Test_Result.xlsx')


def write_header():
    workbook = openpyxl.Workbook()
    worksheet2 = workbook.create_sheet('Details')
    worksheet2.cell(1, 1, "SN")
    worksheet2.cell(1, 2, "Test_Summary")
    worksheet2.cell(1, 1, "Result")
    worksheet2.cell(1, 4, "Remarks")
    workbook.save('Test_Result/Automation_Test_Result.xlsx')


def write_summary():
    workbook = openpyxl.load_workbook('Test_Result/Test_Summary.xlsx')
    worksheet = workbook.create_sheet('Summary')
    worksheet.cell(row=1, column=1).value = "Test Executed On"
    worksheet.cell(row=1, column=2).value = time
    worksheet.cell(row=2, column=1).value = "Total Number of Test Case"
    worksheet.cell(row=2, column=2).value = "=CountA(Details!A:A)"
    worksheet.cell(row=3, column=1).value = "Number of Pass Test Case"
    worksheet.cell(row=3, column=2).value = "=CountA(Details!A:A)"
    worksheet.cell(row=4, column=1).value = "Number of Failed Test Case"
    worksheet.cell(row=4, column=2).value = "=CountA(Details!A:A)"


def wait(value):
    try:
        time.sleep(value)
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def open_browser(value):
    try:
        global driver
        if value == 'chrome':
            s = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=s)
            driver.maximize_window()
            result = "PASS"
            remarks = ""
        # Multiple Browser
        elif value == 'firefox':
            driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())
            driver.maximize_window()
            result = "PASS"
            remarks = ""
        # elif value == 'edge':
        #     driver = webdriver.Edge(EdgeChromiumDriverManager().install())
        #     driver.maximize_window()
        #     result = "PASS"
        #     remarks = ""
        #
        # elif value == 'brave':
        #     driver = webdriver.Chrome(ChromeDriverManager(chrome_type=ChromeType.BRAVE).install())
        #     driver.maximize_window()
        #     result = "PASS"
        #     remarks = ""
        #
        # elif value == 'opera':
        #     driver = webdriver.Opera(executable_path=OperaDriverManager().install())
        #     driver.maximize_window()
        #     result = "PASS"
        #     remarks = ""

        else:
            result = "FAIL"
            remarks = (value, "Browser Not Supported")
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def open_url(value):
    try:
        driver.get(value)
        driver.maximize_window()
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def hover(xpath1, xpath2):
    try:
        parent = driver.find_element(By.XPATH, "xpath1")
        child = driver.find_element(By.XPATH, "xpath2")

        act = ActionChains(driver)
        act.move_to_element(parent).move_to_element(child).click().perform()
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


# new tab ko
def click(xpath):
    try:
        driver.find_element(By.XPATH, xpath).click()
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def new_tab(xpath):
    try:
        #  multiple tab
        click_path = driver.find_element(By.XPATH, xpath)
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL).click(click_path).key_up(Keys.CONTROL).perform()
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(4)
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


# def alert(value):
#     try:
#         if value == 'accept':
#             driver.switch_to.alert.accept()
#             result = "PASS"
#             remarks = ""
#         elif value == 'reject':
#             driver.switch_to.alert.dismiss()
#             result = "PASS"
#             remarks = ""
#     except Exception as ex:
#         result = "FAIL"
#         remarks = ex
#     return result, remarks


def input_text(xpath, value):
    try:
        driver.find_element(By.XPATH, xpath).send_keys(value)
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def date_picker(xpath, value):
    try:
        # 1 case simply push date
        driver.find_element(By.XPATH, xpath).send_keys(value)

        # Incomplete
        # 2 case: Choose date from sliding tabular calender
        # day = value.split("/")[0]
        # month = value.split("/")[1]
        # year = value.split("/")[2]

        # Passing static value
        # year = '2022'
        # month ='June'
        # date ='15'
        #
        # driver.find_element(By.XPATH, "date picker xpath") #open datepicker
        #
        # while True:
        #     mon = driver.find_element(By.XPATH,"month value xpath").text
        #     yr = driver.find_element(By.XPATH,"year xpath").text
        #
        #     if mon == month and yr == year:
        #         break
        #     else:
        #         driver.find_element(By.XPATH, "next arrow xpath").click() #Next Arrow
        #       # driver.find_element(By.XPATH. "previous arrow xpath").click() #previous arrow only activating one
        # #select date if only m and y matches further
        # dates = driver.find_element(By.XPATH, "days xpath")
        # for dte in dates:
        #     if dte.text==date:
        #         dte.click()

        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def compare_text(xpath, value):
    try:
        actual_text = driver.find_element(By.XPATH, xpath).text
        try:
            assert actual_text == value
        except AssertionError:
            result = "FAIL"
            remarks = ("Actual value is", actual_text, "Expected value is", value)
        else:
            result = "PASS"
            remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def select_dropdown(xpath, value):
    try:
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def close_browser():
    try:
        driver.quit()
        result = "PASS"
        remarks = ""
    except Exception as ex:
        result = "FAIL"
        remarks = ex
    return result, remarks


def send_selenium_report():
   fromaddr = "whomanishofficial@gmail.com"
   toaddr = "hymanishchauhan@gmail.com"
   password = 'lhrtjxitrmlkslvw'
   #Email Sending Part
   msg = MIMEMultipart()
   msg['From'] = fromaddr
   msg['To'] = toaddr
   msg['Subject'] = "MeroLagani Test_Case"
   body = MIMEText("Hi , <br> <br> Test for successful transfer <br> <br> Thank You", 'html', 'utf-8')
   msg.attach(body)  # add message body (text or html)
   yourpath = 'test_case/merolagani_testcase.xlsx'  # path for test result or files
   for subdir, dirs, files in os.walk(yourpath):
       for filename in files:
           attachment = open(filename, "rb")
           part = MIMEBase('application', 'octet-stream')
           part.set_payload((attachment).read())
           encoders.encode_base64(part)
           part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
           msg.attach(part)
   server = smtplib.SMTP('smtp.gmail.com', 587)
   server.starttls()
   server.login(fromaddr, password)
   text = msg.as_string()
   server.sendmail(fromaddr, toaddr, text)
   print("Email Sent Sucessfully")
   server.quit()


if __name__ == "__main__":
    read_excel()
    send_selenium_report()

